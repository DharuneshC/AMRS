!pip install ipywidgets
!pip install paddleocr
!pip install paddlepaddle
!pip install opencv-python
!pip install matplotlib
!pip install tabulate
!pip install pandas openpyxl

# Step 2: Import libraries
from paddleocr import PaddleOCR
import cv2
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
import openpyxl
import os
from PIL import Image
import ipywidgets as widgets
from IPython.display import display, clear_output
import re  # ✅ For digit extraction

# Step 3: Global variables
result = ""
extracted_digits = []

# Step 4: OCR Processing + Excel Update
def process_image(image_path):
    global result, extracted_digits
    extracted_digits = []

    # Load image
    img = cv2.imread(image_path)
    img_rgb = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
    plt.figure(figsize=(10, 6))
    plt.imshow(img_rgb)
    plt.title("Input Image")
    plt.axis('off')
    plt.show()

    # Run OCR
    ocr = PaddleOCR(use_angle_cls=True, lang='en')
    result_ocr = ocr.ocr(image_path, cls=True)

    # Extract digits from OCR results
    confidence_threshold = 0.5
    for line in result_ocr[0]:
        box, (text, confidence) = line
        if confidence < confidence_threshold or text.strip() == "":
            continue
        digits_found = re.findall(r'\d+', text)
        extracted_digits.extend(digits_found)

    # ✅ Update Excel based on digit list
    try:
        excel_path = "/content/Book1.xlsx"
        wb = openpyxl.load_workbook(excel_path)
        sheet = wb.active
        row = 3  # Change if needed

        if len(extracted_digits) >= 33:
            # 1. SERIAL_NO in Column A
            sheet.cell(row=row, column=1, value=int(extracted_digits[0]))

            # 2. Skip [1:11], then insert [11:18] → Columns B to H
            for i, val in enumerate(extracted_digits[11:18]):
                sheet.cell(row=row, column=i + 2, value=int(val))

            # 3. Skip [18:28], then insert [28:33] → Columns I to M
            for i, val in enumerate(extracted_digits[28:33]):
                sheet.cell(row=row, column=i + 9, value=int(val))

            wb.save(excel_path)
            result = "✅ Excel updated with selected OCR digits only.\n"
        else:
            result = "⚠️ Not enough digits extracted to update Excel properly.\n"

    except Exception as e:
        result = f"❌ Failed to update Excel: {e}\n"

    with output:
        clear_output()
        print("Extracted Digits:", extracted_digits)
        print(result)

# Step 5: GUI Setup
output = widgets.Output()

upload_button = widgets.FileUpload(
    accept='image/*',
    multiple=False
)
display(upload_button)

process_button = widgets.Button(description="Process Image and Update Excel")
display(process_button, output)

def on_process_button_clicked(b):
    uploaded_file = upload_button.data[0]
    with open("uploaded_image.jpg", "wb") as f:
        f.write(uploaded_file)
    process_image("uploaded_image.jpg")

process_button.on_click(on_process_button_clicked)
